import pandas as pd
from playwright.sync_api import sync_playwright
import re
import time
import json
import os
import random
import math

CACHE_FILE = "processed_links.json"

def load_cache():
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE, "r") as f:
            return set(json.load(f))
    return set()

def save_cache(cache_set):
    with open(CACHE_FILE, "w") as f:
        json.dump(list(cache_set), f, indent=4)

# ==========================================
# HELPER GERAKAN MOUSE NATIVE PLAYWRIGHT
# ==========================================

def human_move_mouse(page, start_x, start_y, end_x, end_y, steps=8):
    """
    Menggerakkan kursor mouse virtual dengan lintasan halus 
    tanpa membanjiri event browser (natural steps).
    """
    mid_x = (start_x + end_x) / 2
    mid_y = (start_y + end_y) / 2
    offset = random.uniform(-30, 30)
    ctrl_x = mid_x + offset
    ctrl_y = mid_y - offset

    for i in range(1, steps + 1):
        t = i / steps
        x = (1 - t) ** 2 * start_x + 2 * (1 - t) * t * ctrl_x + t ** 2 * end_x
        y = (1 - t) ** 2 * start_y + 2 * (1 - t) * t * ctrl_y + t ** 2 * end_y
        
        page.mouse.move(x, y)
        time.sleep(random.uniform(0.01, 0.03))

def human_click_box(page, box, current_pos):
    """
    Menggerakkan kursor dari `current_pos` ke elemen target (`box`),
    lalu mengkliknya secara fisik menggunakan `page.mouse.click`.
    """
    if not box:
        return False
    
    # Ambil titik acak di dalam area elemen (tidak selalu persis di tengah)
    padding_x = box['width'] * 0.25
    padding_y = box['height'] * 0.25
    target_x = box['x'] + random.uniform(padding_x, max(padding_x, box['width'] - padding_x))
    target_y = box['y'] + random.uniform(padding_y, max(padding_y, box['height'] - padding_y))

    start_x = current_pos.get('x', 500)
    start_y = current_pos.get('y', 500)

    # 1. Gerakkan kursor ke target
    human_move_mouse(page, start_x, start_y, target_x, target_y, steps=random.randint(6, 10))

    # 2. Jeda mikro alami manusia sebelum menekan mouse (100-250ms)
    time.sleep(random.uniform(0.1, 0.25))

    # 3. Klik fisik native Playwright (isTrusted=true)
    page.mouse.click(target_x, target_y)

    # Update posisi kursor saat ini
    current_pos['x'] = target_x
    current_pos['y'] = target_y
    return True

# ==========================================

def check_is_bot_or_blocked(page):
    """Mengecek apakah halaman saat ini menunjukkan pesan terdeteksi bot, WAF, atau error SSO."""
    try:
        if page.locator("text=/mendeteksi koneksi anda sebagai bot/i").count() > 0:
            return True
        if page.locator("text=/HaloSIS/i").count() > 0:
            return True
        if page.locator("text=/Lanjutkan dengan SSO/i").count() > 0:
            return True
        if page.locator("text=/Access Denied/i").count() > 0:
            return True
        url = page.url.lower()
        if "sso" in url and ("error" in url or "login" in url or "block" in url):
            return True
    except:
        pass
    return False

def resolve_bot_detection(page, target_link, current_pos=None):
    """
    Penanganan terdeteksi bot (WAF/SSO):
    1. Tunggu beberapa saat (5 detik)
    2. Refresh/reload halaman
    3. Cek & klik tombol 'Lanjutkan dengan SSO' jika ada
    4. Ulangi otomatis 3 kali sebelum minta intervensi manual.
    """
    if not check_is_bot_or_blocked(page):
        return True

    print("="*60)
    print(" [WAF / BOT DETECTED] Halaman terdeteksi bot atau terganggu sesi SSO.")
    print(" Memulai prosedur pemulihan otomatis...")
    
    max_retries = 3
    for attempt in range(1, max_retries + 1):
        print(f" -> [Percobaan {attempt}/{max_retries}] Menunggu 5 detik...")
        time.sleep(5)
        
        print(" -> Refresh/reload halaman...")
        try:
            page.reload(wait_until="networkidle", timeout=15000)
        except Exception:
            try:
                page.goto(target_link, wait_until="networkidle", timeout=15000)
            except Exception:
                pass
        time.sleep(3)
        
        # Cari dan klik tombol 'Lanjutkan dengan SSO' jika ada
        try:
            sso_btn = page.locator("button, a, div").filter(
                has_text=re.compile(r"Lanjutkan dengan SSO|Lanjutkan SSO|Login.*SSO|Masuk.*SSO|Lanjutkan", re.IGNORECASE)
            )
            if sso_btn.count() > 0 and sso_btn.first.is_visible():
                print(" -> Mengklik tombol 'Lanjutkan dengan SSO'...")
                box = sso_btn.first.bounding_box()
                if box and current_pos is not None:
                    human_click_box(page, box, current_pos)
                else:
                    sso_btn.first.click()
                time.sleep(4)
                try:
                    page.wait_for_load_state("networkidle", timeout=10000)
                except Exception:
                    pass
        except Exception as e:
            print(f" -> Cek tombol SSO info: {e}")

        if not check_is_bot_or_blocked(page):
            print(" -> [BERHASIL] Halaman terbebas dari deteksi bot! Melanjutkan bot...")
            print("="*60)
            return True

    print("="*60)
    print(" TERDETEKSI SEBAGAI BOT OLEH SERVER (WAF BPS)!")
    print(" Pemulihan otomatis belum berhasil. Silakan selesaikan di browser.")
    print(" Tekan ENTER di terminal ini jika halaman sudah kembali normal.")
    print("="*60)
    input("Tekan ENTER untuk melanjutkan bot...")
    try:
        page.goto(target_link, wait_until="networkidle", timeout=15000)
    except Exception:
        pass
    return True

def main():
    data_dir = os.path.join(os.getcwd(), "data")
    
    if not os.path.exists(data_dir):
        os.makedirs(data_dir)
        print(f"Folder 'data' telah dibuat di {data_dir}")
        print("Silakan masukkan file Excel anomali ke dalam folder tersebut lalu jalankan ulang script.")
        return
        
    excel_files = [f for f in os.listdir(data_dir) if f.endswith(".xlsx")]
    if not excel_files:
        print("TIDAK ADA FILE EXCEL DITEMUKAN!")
        print(f"Silakan masukkan file Excel anomali (.xlsx) ke dalam folder: {data_dir}")
        return

    raw_links = []
    import openpyxl

    for file_name in excel_files:
        excel_file = os.path.join(data_dir, file_name)
        print(f"Membaca file Excel: {file_name}...")
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
            file_links = []
            for row in range(1, ws.max_row + 1):
                if ws.row_dimensions[row].hidden:
                    continue
                
                status_val = ws.cell(row=row, column=15).value
                if status_val is None or "belum ditindaklanjuti" not in str(status_val).strip().lower():
                    continue

                val = ws.cell(row=row, column=18).value
                if val is not None and str(val).strip() != "":
                    file_links.append(str(val).strip())
            wb.close()
            print(f"  -> {len(file_links)} link dibaca dari {file_name} (hanya status 'Belum Ditindaklanjuti' & visible).")
            raw_links.extend(file_links)
        except Exception as e:
            print(f"  -> Gagal membaca file {file_name}: {e}")

    if not raw_links:
        print("Tidak ada data yang berhasil dibaca dari file Excel mana pun.")
        return

    kegiatan_id = "fd68e454-ba45-4b85-8205-f3bf777ded24"
    edit_links = []
    
    for link in raw_links:
        link = link.strip()
        match = re.search(r"assignment-detail/([a-zA-Z0-9\-]+)", link)
        if match:
            dynamic_id = match.group(1)
            new_link = f"https://fasih-sm.bps.go.id/app/assignment/{kegiatan_id}/{dynamic_id}/edit"
            edit_links.append(new_link)

    print(f"Ditemukan {len(edit_links)} link anomali dalam file Excel.")

    if not edit_links:
        print("Tidak ada link valid yang ditemukan di file Excel.")
        return

    processed_cache = load_cache()
    pending_links = [link for link in edit_links if link not in processed_cache]
    
    print(f"Total link: {len(edit_links)}")
    print(f"Sudah diproses (dari cache): {len(processed_cache)}")
    print(f"Sisa yang akan diproses: {len(pending_links)}")
    
    if not pending_links:
        print("Semua data untuk kodekec ini sudah berhasil diproses!")
        return

    print("="*60)
    print("Membuka browser Playwright...")
    
    with sync_playwright() as p:
        user_data_dir = os.path.join(os.getcwd(), "chrome_profile_anomali")
        
        # Opsi 1: Coba gunakan Google Chrome asli jika terpasang (jauh lebih aman dari WAF)
        launch_args = [
            "--disable-blink-features=AutomationControlled",
            "--start-maximized"
        ]
        
        context_kwargs = {
            "user_data_dir": user_data_dir,
            "headless": False,
            "args": launch_args,
            "user_agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "no_viewport": True
        }
        
        # Coba buka via channel 'chrome' (Google Chrome asli pengguna)
        try:
            context = p.chromium.launch_persistent_context(channel="chrome", **context_kwargs)
            print("  -> Menjalankan via Google Chrome Asli (System Chrome)...")
        except Exception:
            context = p.chromium.launch_persistent_context(**context_kwargs)
            print("  -> Menjalankan via Playwright Chromium...")
        
        page = context.pages[0] if context.pages else context.new_page()
        
        # Bypass webdriver sederhana (aman tanpa jebakan Object.defineProperty)
        page.add_init_script("delete navigator.__proto__.webdriver;")

        current_pos = {"x": random.randint(300, 700), "y": random.randint(200, 500)}

        page.goto("https://fasih-sm.bps.go.id/")
        
        print("Silakan login ke Fasih-SM (SSO BPS) jika belum login.")
        print("="*60)
        input("Tekan ENTER di terminal ini jika sudah berhasil login dan siap memulai...")

        for idx, link in enumerate(edit_links, 1):
            if link in processed_cache:
                print(f"[{idx}/{len(edit_links)}] SKIP (Sudah diproses): {link}")
                continue
                
            print(f"[{idx}/{len(edit_links)}] Memproses: {link}")
            try:
                # Navigasi ke halaman EDIT
                try:
                    page.goto(link, wait_until="networkidle", timeout=30000)
                except Exception as e:
                    if "interrupted by another navigation" in str(e):
                        print("  -> [Info] Navigasi dialihkan oleh SSO Refresh. Mencoba ulang...")
                        time.sleep(3)
                        try:
                            page.goto(link, wait_until="networkidle", timeout=30000)
                        except Exception:
                            pass
                    else:
                        raise e
                
                # Cek dan tangani jika terkena blokir bot (WAF/SSO)
                resolve_bot_detection(page, link, current_pos)
                
                # Verifikasi bahwa browser benar-benar berada di halaman /edit
                current_url = page.url
                if "/edit" in current_url:
                    on_edit_page = True
                else:
                    print(f"  -> [SKIP] Browser tidak di halaman /edit (URL: {current_url}). Bukan wilayah admin, langsung skip.")
                    processed_cache.add(link)
                    save_cache(processed_cache)
                    continue
                
                # 1. Masuk ke menu Catatan
                print("  -> Mengarahkan mouse ke tab 'Catatan'...")
                tab_catatan = page.get_by_role("tab", name="Catatan", exact=False)
                if tab_catatan.count() == 0:
                    tab_catatan = page.get_by_text("Catatan", exact=False).first
                
                box_catatan = None
                try:
                    tab_catatan.first.wait_for(state="visible", timeout=10000)
                    box_catatan = tab_catatan.first.bounding_box()
                except Exception:
                    pass
                
                if box_catatan:
                    human_click_box(page, box_catatan, current_pos)
                else:
                    tab_catatan.first.click(timeout=10000)
                time.sleep(1.2)
                
                # 2. Centang checkbox "Tampilkan Anomali Usaha dan Keluarga"
                print("  -> Mencari dan mengklik checkbox 'Tampilkan Anomali'...")
                cb_info = page.evaluate('''() => {
                    const elements = Array.from(document.querySelectorAll('*'));
                    const targetEl = elements.find(el => el.textContent && el.textContent.toLowerCase().includes('tampilkan anomali') && el.children.length === 0);
                    
                    if (targetEl) {
                        let parent = targetEl.parentElement;
                        for(let i=0; i<5; i++) {
                            if(!parent) break;
                            const cb = parent.querySelector('input[type="checkbox"]');
                            if(cb) {
                                const rect = cb.getBoundingClientRect();
                                return { 
                                    found: true, 
                                    was_checked: cb.checked, 
                                    rect: { x: rect.x, y: rect.y, width: rect.width, height: rect.height } 
                                };
                            }
                            parent = parent.parentElement;
                        }
                        const rect = targetEl.getBoundingClientRect();
                        return { 
                            found: true, 
                            was_checked: false, 
                            rect: { x: rect.x, y: rect.y, width: rect.width, height: rect.height } 
                        };
                    }
                    return { found: false, was_checked: false };
                }''')
                
                if cb_info and cb_info.get('found'):
                    if cb_info.get('was_checked'):
                        print("  -> Checkbox sudah tercentang dari awal. Melewati klik 'Kirim'.")
                    else:
                        rect = cb_info.get('rect')
                        if rect and rect.get('width', 0) > 0:
                            human_click_box(page, rect, current_pos)
                        else:
                            page.locator("text=/Tampilkan Anomali/i").first.click(force=True)
                        time.sleep(1.2)

                        # 3. Klik Kirim (Tombol utama)
                        print("  -> Mengarahkan mouse ke tombol 'Kirim'...")
                        kirim_clicked = False
                        
                        kirim_rect = page.evaluate('''() => {
                            const btns = Array.from(document.querySelectorAll('button'));
                            const target = btns.find(b => {
                                const t = (b.textContent || '').toLowerCase().trim();
                                return (t.includes('kirim') || t.includes('submit') || t.includes('simpan')) && b.offsetParent !== null;
                            });
                            if (target) {
                                const rect = target.getBoundingClientRect();
                                return { x: rect.x, y: rect.y, width: rect.width, height: rect.height };
                            }
                            return null;
                        }''')
                        
                        if kirim_rect and kirim_rect.get('width', 0) > 0:
                            human_click_box(page, kirim_rect, current_pos)
                            kirim_clicked = True
                        else:
                            try:
                                kirim_btn = page.get_by_role("button", name="Kirim").first
                                if kirim_btn.is_visible():
                                    box = kirim_btn.bounding_box()
                                    if box:
                                        human_click_box(page, box, current_pos)
                                    else:
                                        kirim_btn.click()
                                    kirim_clicked = True
                            except Exception:
                                pass

                        if kirim_clicked:
                            time.sleep(1.5)
                            
                            # 4. Handle Konfirmasi Modal
                            for _ in range(3):
                                confirm_rect = page.evaluate('''() => {
                                    const modal = document.querySelector('.modal, [role="dialog"], [data-state="open"]');
                                    if (modal) {
                                        const modalBtns = Array.from(modal.querySelectorAll('button'));
                                        if (modalBtns.length > 0) {
                                            const target = modalBtns[modalBtns.length - 1];
                                            const rect = target.getBoundingClientRect();
                                            return { x: rect.x, y: rect.y, width: rect.width, height: rect.height };
                                        }
                                    }
                                    const btns = Array.from(document.querySelectorAll('button'));
                                    const target = btns.find(b => {
                                        const t = (b.textContent || '').toLowerCase().trim();
                                        return (t.includes('konfirmasi') || t.includes('ya') || t.includes('setuju')) && b.offsetParent !== null;
                                    });
                                    if (target) {
                                        const rect = target.getBoundingClientRect();
                                        return { x: rect.x, y: rect.y, width: rect.width, height: rect.height };
                                    }
                                    return null;
                                }''')
                                
                                if confirm_rect and confirm_rect.get('width', 0) > 0:
                                    print("     -> Mengklik tombol konfirmasi dialog...")
                                    human_click_box(page, confirm_rect, current_pos)
                                    time.sleep(1.5)
                                else:
                                    break
                            
                            time.sleep(2)
                        else:
                            print("  -> [Warning] Tombol 'Kirim' tidak ditemukan. Lanjut ke Reject...")
                else:
                    try:
                        print("  -> (Fallback) Mencoba klik paksa teks 'Tampilkan Anomali'...")
                        el = page.locator("text=/Tampilkan Anomali/i").first
                        box = el.bounding_box()
                        if box:
                            human_click_box(page, box, current_pos)
                        else:
                            el.click(force=True)
                        time.sleep(1)
                    except Exception:
                        pass
                
                # --- FASE 2: REJECT ---
                base_link = link.replace("/edit", "")
                print(f"  -> Membuka halaman detail untuk Reject: {base_link}")
                
                try:
                    page.goto(base_link, wait_until="networkidle")
                except Exception as e:
                    if "interrupted by another navigation" in str(e):
                        time.sleep(2)
                        page.goto(base_link, wait_until="networkidle")
                    else:
                        raise e
                        
                time.sleep(2)
                resolve_bot_detection(page, base_link, current_pos)
                
                # Klik tombol Reject atau X
                print("  -> Mengarahkan mouse ke tombol Reject / X...")
                reject_success = False
                for _ in range(10):
                    reject_rect = page.evaluate('''() => {
                        const buttons = Array.from(document.querySelectorAll('button'));
                        
                        let target = buttons.find(b => {
                            const t = (b.textContent || '').toLowerCase().trim();
                            const a = (b.getAttribute('aria-label') || '').toLowerCase();
                            const title = (b.getAttribute('title') || '').toLowerCase();
                            return t.includes('reject') || t.includes('tolak') || t === 'x' || t === '×' || 
                                   a.includes('reject') || a.includes('tolak') || title.includes('reject');
                        });
                        
                        if (!target) {
                            const dangerBtns = buttons.filter(b => {
                                const c = (b.className || '').toLowerCase();
                                return c.includes('reject') || c.includes('danger') || c.includes('red-') || c.includes('destructive');
                            });
                            if (dangerBtns.length > 0) {
                                target = dangerBtns.reduce((prev, curr) => {
                                    return (curr.getBoundingClientRect().bottom > prev.getBoundingClientRect().bottom) ? curr : prev;
                                });
                            }
                        }
                        
                        if (!target) {
                            target = buttons.find(b => {
                                const rect = b.getBoundingClientRect();
                                return rect.left > window.innerWidth * 0.7 && rect.top > window.innerHeight * 0.7;
                            });
                        }
                        
                        if (target && target.offsetParent !== null) {
                            const rect = target.getBoundingClientRect();
                            return { x: rect.x, y: rect.y, width: rect.width, height: rect.height };
                        }
                        return null;
                    }''')
                    
                    if reject_rect and reject_rect.get('width', 0) > 0:
                        human_click_box(page, reject_rect, current_pos)
                        reject_success = True
                        break
                    time.sleep(1)
                
                if not reject_success:
                    raise Exception("Gagal menemukan tombol Reject/X di halaman setelah menunggu 10 detik.")
                
                time.sleep(1.5)
                
                # Klik tombol Konfirmasi di dialog
                print("  -> Mengarahkan mouse ke tombol Konfirmasi...")
                confirm_success = False
                for _ in range(10):
                    confirm_rect = page.evaluate('''() => {
                        const buttons = Array.from(document.querySelectorAll('button'));
                        let target = buttons.find(b => {
                            const t = (b.textContent || '').toLowerCase().trim();
                            return t.includes('konfirmasi') || t.includes('ya') || t.includes('setuju') || t.includes('lanjut');
                        });
                        
                        if (!target) {
                            const modal = document.querySelector('.modal, [role="dialog"], [data-state="open"]');
                            if (modal) {
                                const modalBtns = Array.from(modal.querySelectorAll('button'));
                                if (modalBtns.length > 0) {
                                    target = modalBtns[modalBtns.length - 1];
                                }
                            }
                        }
                        
                        if (target && target.offsetParent !== null) {
                            const rect = target.getBoundingClientRect();
                            return { x: rect.x, y: rect.y, width: rect.width, height: rect.height };
                        }
                        return null;
                    }''')
                    
                    if confirm_rect and confirm_rect.get('width', 0) > 0:
                        human_click_box(page, confirm_rect, current_pos)
                        confirm_success = True
                        break
                    time.sleep(1)
                
                if not confirm_success:
                    print("     [Warning] Tombol Konfirmasi tidak ditemukan. Melanjutkan...")
                
                time.sleep(2.5)
                
                processed_cache.add(link)
                save_cache(processed_cache)
                print(f"  -> Sukses dikirim, direject, dan disimpan ke cache.")
                
            except Exception as e:
                print(f"  -> Terjadi error pada link {link}:")
                print(f"     {e}")
                if check_is_bot_or_blocked(page):
                    resolve_bot_detection(page, link, current_pos)
                print("     Lanjut ke link berikutnya...")

            # Delay alami 4-8 detik antar link (tanpa spam event mouse)
            delay = random.uniform(4.0, 8.0)
            time.sleep(delay)

        print("Proses otomatisasi selesai!")
        context.close()

if __name__ == "__main__":
    main()
